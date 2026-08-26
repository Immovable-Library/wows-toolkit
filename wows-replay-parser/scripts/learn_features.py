#!/usr/bin/env python3
"""Self-learning sync: 备注 -> 特色列 + table style normalization.

Reads reports/舰船特色表.xlsx and for every row with a 备注:

1. matches existing feature columns via a keyword table (and by full name for
   learned columns) and fills "1" where the note mentions them, consuming the
   matched text;
2. extracts remaining feature phrases and either merges them into an existing
   column (containment) or adds them as NEW feature columns before 备注.

Weapon-context handling: when a note is about torpedoes (contains 鱼雷/雷),
generic modifiers such as 快装填 / 高标伤 are qualified as torpedo features
(鱼雷装填快 / 鱼雷高标伤) instead of generic ship traits, and are never
merged into main-battery columns (高单轮标伤).

Also normalizes the table for WPS/Excel:
- freeze header row + ship-name columns (freeze panes at F2);
- auto-fit column widths from content (CJK counts as 2 chars);
- keeps the autoFilter range but clears active filter criteria.

Run after the user annotates and before feature_gate / analysis:
  python scripts/learn_features.py
"""
from __future__ import annotations

import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "reports" / "舰船特色表.xlsx"
SHEET = "舰船特色表"

# keyword table for the base 28 feature columns (variants allowed)
KEYWORDS = {
    "大口径主炮": ["大口径主炮", "大口径", "口径大", "大管子"],
    "高DPM/高射速": ["高DPM/高射速", "高DPM", "高射速", "高dpm", "dpm高"],
    "高单轮标伤": ["高单轮标伤", "单轮标伤", "高标伤", "标伤高", "标伤"],
    "AP穿深/碾压特化": ["AP穿深/碾压特化", "ap穿深", "穿深高", "碾压", "高穿深"],
    "高精度主炮": ["高精度主炮", "高精度", "精度高", "散布好", "精度"],
    "大射程/远程狙击": ["大射程/远程狙击", "大射程", "远程狙击", "射程远", "超远射程"],
    "点火流": ["点火流", "点火高", "高点火", "起火", "点火"],
    "齐射角好": ["齐射角好", "齐射角", "射角好", "射界好"],
    "强鱼雷": ["强鱼雷", "鱼雷强", "高伤鱼雷", "强力鱼雷"],
    "隐蔽鱼雷": ["隐蔽鱼雷", "鱼雷隐蔽", "低发现鱼雷", "雷隐蔽"],
    "鱼雷装填手": ["鱼雷装填手", "装填手", "鱼雷装填手机制"],
    "副炮流": ["副炮流", "副炮强", "强副炮", "副炮"],
    "手动副炮": ["手动副炮", "手动控制副炮", "手动副炮组"],
    "强防空": ["强防空", "防空强", "防空好", "防空"],
    "航空/空袭": ["航空/空袭", "航空", "空袭", "航母", "飞机"],
    "侦察/战斗机": ["侦察/战斗机", "侦察机", "战斗机", "侦察"],
    "高机动": ["高机动", "机动好", "航速高", "航速快", "速度快", "加速快", "高航速"],
    "灵活转向": ["灵活转向", "转向好", "转舵快", "转向半径小", "转舵"],
    "高生存/大修": ["高生存/大修", "高生存", "生存强", "大修", "维修小组", "血厚", "血量高", "高血量"],
    "鱼雷防护好": ["鱼雷防护好", "鱼雷防护", "抗雷"],
    "优秀隐蔽": ["优秀隐蔽", "隐蔽好", "隐蔽优秀", "隐蔽低", "隐蔽"],
    "烟内开火惩罚小": ["烟内开火惩罚小", "烟内开火", "烟中开火"],
    "水听": ["水听", "水听器", "水底搜索"],
    "雷达": ["雷达"],
    "烟幕": ["烟幕", "烟雾", "烟雾发生器", "发烟机"],
    "装填助推爆发": ["装填助推爆发", "装填助推", "助推器", "主炮装填助推", "鱼雷装填助推", "爆发装填"],
    "反潜": ["反潜", "深水炸弹", "反潜空袭"],
    "潜艇/深潜": ["潜艇/深潜", "潜艇", "深潜"],
}

# keywords that are ambiguous when the note talks about torpedoes
TORPEDO_AMBIGUOUS = {"标伤", "高标伤", "单轮标伤"}

# canonical names for torpedo-qualified modifiers
TORPEDO_CANON = {
    "快装填": "鱼雷装填快",
    "装填快": "鱼雷装填快",
    "装填较快": "鱼雷装填快",
    "装填较": "鱼雷装填快",
    "较快装填": "鱼雷装填快",
    "高标伤": "鱼雷高标伤",
    "标伤高": "鱼雷高标伤",
    "高伤雷": "鱼雷高标伤",
    "自爆雷": "短程自爆雷",
    "短程自爆雷": "短程自爆雷",
}


def canon_torpedo(ph: str) -> str:
    core = ph[2:] if ph.startswith("鱼雷") else ph
    return TORPEDO_CANON.get(core, ph)

SEP = re.compile(r"[；;，,、/和及与\s()（）\[\]【】]+")
SPLIT_STARTERS = ["短程", "自爆", "快速", "快", "高", "强", "远", "近", "低", "大", "小", "主", "副", "鱼雷"]
MERGE_MODIFIERS = ("短程", "自爆", "高", "强", "快", "远", "近", "低", "大", "小", "主", "副")
STOP = {"较快", "很快", "比较", "非常", "主要", "特色", "属于", "擅长", "适合", "就是", "可以", "因为", "所以", "有点", "一个"}


def cjk_len(s: str) -> int:
    return sum(1 for ch in s if "\u4e00" <= ch <= "\u9fff")


def clean_phrase(s: str) -> str:
    s = s.strip().strip("：:。.！!？?的型流")
    s = re.sub(r"^(有|带|是|属于|主打|擅长|适合|比较|非常|有点)+", "", s)
    s = re.sub(r"^(?:而|且|并|也|又|还|更)+", "", s)
    return s.strip()


def split_phrase(phrase: str):
    if cjk_len(phrase) <= 4:
        return [phrase]
    pat = "(?=" + "|".join(re.escape(w) for w in SPLIT_STARTERS) + ")"
    return [p for p in re.split(pat, phrase) if cjk_len(p) >= 2]


def extract_candidates(note: str, torpedo_ctx: bool):
    cands = []
    for raw in SEP.split(note):
        if not raw:
            continue
        for ph in split_phrase(raw):
            ph = clean_phrase(ph)
            if cjk_len(ph) < 2 or ph in STOP:
                continue
            if torpedo_ctx:
                ph = canon_torpedo(ph)
            if cands and cjk_len(cands[-1]) <= 2 and ph.startswith(MERGE_MODIFIERS):
                cands[-1] = cands[-1] + ph
            else:
                cands.append(ph)
    return cands


def display_width(s) -> int:
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(s or ""))


def main():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    header = [c.value for c in ws[1]]
    try:
        note_idx = header.index("备注")
    except ValueError:
        note_idx = len(header) - 1
    try:
        nation_idx = header.index("国家")
    except ValueError:
        nation_idx = 4
    feat_indices = list(range(nation_idx + 1, note_idx))
    feat_names = [header[i] for i in feat_indices]

    def kw_map_for(torpedo_ctx):
        out = {}
        for name in feat_names:
            if name in KEYWORDS:
                kws = list(KEYWORDS[name])
                if torpedo_ctx and name == "高单轮标伤":
                    kws = [k for k in kws if k not in TORPEDO_AMBIGUOUS]
                out[name] = kws
            else:
                out[name] = [p for p in [name] + (name.split("/") if name else []) if len(p) >= 2]
        return out

    rows = list(ws.iter_rows(min_row=2))
    new_cols = {}
    summary = []

    for row in rows:
        r = row[0].row
        note = row[note_idx].value
        if note is None:
            continue
        note = str(note).strip()
        if not note:
            continue
        torpedo_ctx = ("鱼雷" in note) or ("雷" in note)
        working = note
        marked = []
        kw_map = kw_map_for(torpedo_ctx)
        all_kws = sorted(
            ((kw, name) for name, kws in kw_map.items() for kw in kws),
            key=lambda kv: -len(kv[0]),
        )
        for kw, name in all_kws:
            if len(kw) < 2:
                continue
            if kw in working:
                marked.append(name)
                working = working.replace(kw, "", 1)
        for name in set(marked):
            idx = feat_indices[feat_names.index(name)]
            if ws.cell(row=r, column=idx + 1).value in (None, ""):
                ws.cell(row=r, column=idx + 1, value="1")
        cands = extract_candidates(working, torpedo_ctx)
        added = []
        for c in cands:
            hit = next((n for n in feat_names if n and (c in n or n in c)), None)
            if hit:
                marked.append(hit)
                idx = feat_indices[feat_names.index(hit)]
                if ws.cell(row=r, column=idx + 1).value in (None, ""):
                    ws.cell(row=r, column=idx + 1, value="1")
            else:
                new_cols.setdefault(c, set()).add(r)
                added.append(c)
        if marked:
            summary.append((r, note, sorted(set(marked)), added))

    new_names = sorted(new_cols)
    if new_names:
        ws.insert_cols(note_idx + 1, amount=len(new_names))
        src = ws.cell(row=1, column=feat_indices[0] + 1)
        for j, name in enumerate(new_names):
            hc = ws.cell(row=1, column=note_idx + 1 + j, value=name)
            if src.has_style:
                hc._style = src._style
        for name, rows_set in new_cols.items():
            j = new_names.index(name)
            for r in rows_set:
                ws.cell(row=r, column=note_idx + 1 + j, value="1")

    ws.freeze_panes = "F2"
    try:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
        ws.auto_filter.filterColumn = []
    except Exception:
        pass

    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        hdr = ws.cell(row=1, column=col_idx).value
        widths = [display_width(hdr) + 2]
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            v = row[0].value
            if v not in (None, ""):
                widths.append(display_width(v) + 2)
        cap = 40 if col_idx == ws.max_column else 26
        ws.column_dimensions[col_letter].width = max(6, min(max(widths), cap))

    wb.save(XLSX)
    wb.close()

    print(f"学习完成: 新增特色列 {len(new_names)} 个")
    for name in new_names:
        print(f"  新增列: {name} -> 行 {sorted(new_cols[name])}")
    for r, note, marked, added in summary:
        print(f"  行 {r}: 备注 {note!r} -> 命中 {marked} / 新增 {added}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())